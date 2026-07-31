"""Tests for doc pipeline deep improvements (DC03–DC06).

Tests:
  DC03 — chunk_by_section, chunk_by_paragraph
  DC04 — extract_from_url, _extract_education, _extract_languages,
         education/languages fields in output
  DC05 — process_folder with .pdf, max_files, progress logging
  DC06 — export_to_excel, merge_exports, stats_report
"""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import patch

from app.application.documents.pipeline import (
    _count_filled_fields,
    _extract_education,
    _extract_languages,
    chunk_by_paragraph,
    chunk_by_section,
    chunk_text,
    export_results,
    export_to_excel,
    extract_from_text,
    extract_from_url,
    merge_exports,
    process_folder,
    stats_report,
)


# ══════════════════════════════════════════════════════════════════════════════
# DC03 — chunk_by_section / chunk_by_paragraph
# ══════════════════════════════════════════════════════════════════════════════


def test_chunk_by_section_markdown_headers():
    text = (
        "# Introducción\n"
        "Este es el párrafo introductorio del documento.\n\n"
        "## Metodología\n"
        "La metodología usada es cualitativa con enfoque en entrevistas.\n\n"
        "### Detalles\n"
        "Se realizaron 50 entrevistas a profundidad.\n"
    )
    chunks = chunk_by_section(text, max_chars=500)
    assert len(chunks) >= 1
    # Each chunk should be a dict with required keys
    for c in chunks:
        assert "index" in c
        assert "text" in c
        assert "start_char" in c
        assert "end_char" in c
        assert "char_count" in c
    # First chunk should contain the intro header
    assert any("# Introducción" in c["text"] for c in chunks)


def test_chunk_by_section_empty():
    chunks = chunk_by_section("", max_chars=100)
    assert len(chunks) == 1
    assert chunks[0]["text"] == ""


def test_chunk_by_section_no_headers_falls_back():
    text = "Texto sin headers markdown.\n" * 30
    chunks = chunk_by_section(text, max_chars=200)
    assert len(chunks) >= 1
    # Without headers, sections may be the whole text (chunked by size)
    assert all("index" in c for c in chunks)


def test_chunk_by_section_oversized_section_subdivides():
    """Section larger than max_chars should be subdivided."""
    huge_section = (
        "## Grandes datos\n"
        + ("Datos de prueba que se repiten para llenar espacio. " * 100)
        + "\n"
    )
    chunks = chunk_by_section(huge_section, max_chars=400, overlap=50)
    assert len(chunks) > 1


def test_chunk_by_paragraph_splits_on_blank_lines():
    text = (
        "Párrafo uno con información importante sobre el candidato.\n\n"
        "Párrafo dos con experiencia laboral detallada.\n\n"
        "Párrafo tres con habilidades técnicas y blandas.\n\n"
        "Párrafo cuatro con educación y certificaciones."
    )
    chunks = chunk_by_paragraph(text, max_chars=300)
    assert len(chunks) >= 2
    for c in chunks:
        assert "index" in c
        assert "text" in c
        assert "char_count" in c


def test_chunk_by_paragraph_empty():
    chunks = chunk_by_paragraph("", max_chars=100)
    assert len(chunks) == 1
    assert chunks[0]["text"] == ""


def test_chunk_by_paragraph_single_chunk_when_small():
    text = "Párrafo corto.\n\nOtro párrafo corto."
    chunks = chunk_by_paragraph(text, max_chars=500)
    assert len(chunks) == 1


def test_chunk_by_paragraph_both_modes_same_format():
    """chunk_by_section and chunk_by_paragraph return same shape as chunk_text."""
    text = "Texto de prueba. " * 200

    ct = chunk_text(text, max_chars=300)
    cs = chunk_by_section(text, max_chars=300)
    cp = chunk_by_paragraph(text, max_chars=300)

    for chunks in [ct, cs, cp]:
        assert isinstance(chunks, list)
        for c in chunks:
            for key in ["index", "text", "start_char", "end_char", "char_count"]:
                assert key in c, f"Missing {key} in chunk"


# ══════════════════════════════════════════════════════════════════════════════
# DC04 — Education and languages extraction
# ══════════════════════════════════════════════════════════════════════════════


def test_extract_education_from_section():
    text = (
        "Nombre: Carlos Fuentes\n"
        "Educación:\n"
        "- Licenciatura en Computación — Universidad Central de Venezuela\n"
        "- Maestría en Ciencias de Datos — Universidad Simón Bolívar\n"
        "- Diplomado en Gerencia de Proyectos\n"
    )
    edu = _extract_education(text)
    assert len(edu) >= 2
    assert any("Computación" in e for e in edu)


def test_extract_education_keyword_matching():
    text = (
        "Perfil Profesional\n"
        "Estudios: Ingeniería de Sistemas en la Universidad Nacional.\n"
        "Posee un master en administración de empresas."
    )
    edu = _extract_education(text)
    assert len(edu) >= 1


def test_extract_education_empty_text():
    assert _extract_education("") == []
    assert _extract_education("   ") == []


def test_extract_languages_from_section():
    text = (
        "Idiomas:\n"
        "Inglés avanzado, Español nativo, Francés intermedio\n"
    )
    langs = _extract_languages(text)
    assert len(langs) >= 2
    found_langs = [l.lower() for l in langs]
    assert any("inglés" in l or "ingles" in l for l in found_langs)
    assert any("español" in l or "espanol" in l for l in found_langs)


def test_extract_languages_full_text_fallback():
    text = (
        "Nombre: María González\n"
        "Experiencia: trabajó con clientes de habla inglesa y portuguesa.\n"
        "Maneja el alemán a nivel básico."
    )
    langs = _extract_languages(text)
    assert len(langs) >= 1
    found = [l.lower() for l in langs]
    assert any("inglés" in l or "ingles" in l for l in found)


def test_extract_languages_empty_text():
    assert _extract_languages("") == []
    assert _extract_languages("   ") == []


def test_extraction_includes_education_and_languages_fields():
    text = (
        "Nombre: Ana Pérez\n"
        "Email: ana@ejemplo.com\n"
        "Educación: Ingeniería en Sistemas — UCV\n"
        "Idiomas: Inglés, Francés\n"
    )
    result = extract_from_text(text)
    assert result["ok"] is True
    fields = result["fields"]
    assert "education" in fields
    assert "languages" in fields
    assert isinstance(fields["education"], list)
    assert isinstance(fields["languages"], list)


# ══════════════════════════════════════════════════════════════════════════════
# DC04 — extract_from_url
# ══════════════════════════════════════════════════════════════════════════════


def test_extract_from_url_success(monkeypatch):
    """Simulate a successful HTTP response."""

    class FakeResponse:
        status_code = 200
        text = (
            "<html><body>"
            "<h1>CV de Juan Pérez</h1>"
            "<p>Email: juan@test.com</p>"
            "<p>Tel: +58 412 555 1234</p>"
            "<p>Educación: Ingeniería — UCV</p>"
            "</body></html>"
        )
        headers = {"content-type": "text/html; charset=utf-8"}

    class FakeClient:
        def __init__(self, *args, **kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

        def get(self, url, headers=None):
            return FakeResponse()

    monkeypatch.setattr(
        "app.application.documents.pipeline.httpx.Client",
        FakeClient,
    )

    result = extract_from_url("https://example.com/cv.html")
    assert result["ok"] is True
    assert result["source"] == "url"
    assert result["url"] == "https://example.com/cv.html"
    assert result["status_code"] == 200
    assert "juan@test.com" in result["fields"]["email"]


def test_extract_from_url_timeout(monkeypatch):
    """Simulate a timeout."""

    import httpx

    class FakeClient:
        def __init__(self, *args, **kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

        def get(self, url, headers=None):
            raise httpx.TimeoutException("timeout")

    monkeypatch.setattr(
        "app.application.documents.pipeline.httpx.Client",
        FakeClient,
    )

    result = extract_from_url("https://slow-site.com")
    assert result["ok"] is False
    assert "Timeout" in result["error"]


def test_extract_from_url_non_200(monkeypatch):
    """Simulate a 404 response."""

    class FakeResponse:
        status_code = 404
        text = "Not Found"
        headers = {"content-type": "text/html"}

    class FakeClient:
        def __init__(self, *args, **kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

        def get(self, url, headers=None):
            return FakeResponse()

    monkeypatch.setattr(
        "app.application.documents.pipeline.httpx.Client",
        FakeClient,
    )

    result = extract_from_url("https://example.com/notfound")
    assert result["ok"] is False
    assert "404" in result["error"]


def test_extract_from_url_request_error(monkeypatch):
    """Simulate a connection error."""

    import httpx

    class FakeClient:
        def __init__(self, *args, **kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

        def get(self, url, headers=None):
            raise httpx.RequestError("connection refused")

    monkeypatch.setattr(
        "app.application.documents.pipeline.httpx.Client",
        FakeClient,
    )

    result = extract_from_url("https://nonexistent.example.com")
    assert result["ok"] is False
    assert "Error al descargar URL" in result["error"]


# ══════════════════════════════════════════════════════════════════════════════
# DC05 — process_folder improvements
# ══════════════════════════════════════════════════════════════════════════════


def test_process_folder_max_files(tmp_path: Path):
    """max_files should limit the number of processed files."""
    for i in range(25):
        (tmp_path / f"cv_{i:02d}.txt").write_text(
            f"Nombre: Persona {i}\nEmail: persona{i}@test.com\n",
            encoding="utf-8",
        )

    result = process_folder(tmp_path, max_files=10)
    assert result["ok"] is True
    assert result["total_files"] == 25  # found
    assert result["total_processed"] == 10  # only processed 10
    assert len(result["results"]) <= 10


def test_process_folder_max_files_zero_no_limit(tmp_path: Path):
    """max_files=0 or None should process all."""
    for i in range(5):
        (tmp_path / f"cv_{i}.txt").write_text(
            f"Nombre: Persona {i}\nEmail: p{i}@test.com\n",
            encoding="utf-8",
        )

    result = process_folder(tmp_path, max_files=None)
    assert result["total_processed"] == 5
    assert len(result["results"]) == 5


def test_process_folder_with_pdf_support(tmp_path: Path, monkeypatch):
    """When pdf extraction is available, .pdf files should be included."""
    monkeypatch.setattr(
        "app.application.documents.pipeline._pdf_extraction_available",
        lambda: True,
    )
    monkeypatch.setattr(
        "app.application.documents.pipeline._check_pymupdf",
        lambda: True,
    )
    monkeypatch.setattr(
        "app.application.documents.pipeline._extract_pdf_text",
        lambda path: "Nombre: PDF User\nEmail: pdf@test.com\n",
    )

    (tmp_path / "doc.pdf").write_bytes(b"%PDF-1.4 fake")
    (tmp_path / "doc.txt").write_text("Nombre: TXT User\nEmail: txt@test.com\n", encoding="utf-8")

    result = process_folder(tmp_path)
    assert result["ok"] is True
    assert result["total_files"] == 2
    assert result["total_processed"] == 2
    assert "doc.pdf" in result["results"] or "doc.txt" in result["results"]


def test_process_folder_logs_progress(tmp_path: Path, caplog):
    """Progress should be logged every 10 files."""
    import logging

    caplog.set_level(logging.INFO, logger="dot.doc_pipeline")

    for i in range(25):
        (tmp_path / f"cv_{i:02d}.txt").write_text(
            f"Nombre: Persona {i}\nEmail: p{i}@test.com\n",
            encoding="utf-8",
        )

    result = process_folder(tmp_path)
    assert result["ok"] is True

    progress_logs = [
        r for r in caplog.records
        if "doc_pipeline folder progress" in r.message
    ]
    assert len(progress_logs) >= 2  # at least at 10 and 25

    # Verify the final log has total count
    final_log = progress_logs[-1].message
    assert "25/25" in final_log or "total" in final_log.lower()


def test_process_folder_nonexistent_returns_total_processed():
    result = process_folder("/nonexistent/folder/deep")
    assert result["ok"] is False
    assert result["total_processed"] == 0
    assert result["total_files"] == 0


# ══════════════════════════════════════════════════════════════════════════════
# DC06 — export_to_excel
# ══════════════════════════════════════════════════════════════════════════════


def test_export_to_excel_creates_file(tmp_path: Path):
    """export_to_excel should create a .xlsx file."""
    docs = {
        "cv1.txt": {
            "ok": True,
            "fields": {
                "name": "Ana López",
                "email": "ana@test.com",
                "phone": "555",
                "summary": "Ingeniera",
                "skills": ["python", "react"],
                "experience": ["CEO at X"],
                "education": ["Ingeniería — UCV"],
                "languages": ["Inglés", "Español"],
            },
            "file_type": "txt",
            "path": "/tmp/cv1.txt",
        },
        "cv2.txt": {
            "ok": True,
            "fields": {
                "name": "Luis Gómez",
                "email": "luis@test.com",
                "phone": None,
                "summary": "Dev",
                "skills": ["java"],
                "experience": [],
                "education": [],
                "languages": ["Inglés"],
            },
            "file_type": "txt",
            "path": "/tmp/cv2.txt",
        },
    }

    out = tmp_path / "export.xlsx"
    result_path = export_to_excel(docs, output_path=str(out))

    assert result_path == out
    assert out.is_file()
    assert out.stat().st_size > 0


def test_export_to_excel_with_process_folder_format(tmp_path: Path):
    """Should accept process_folder nested format."""
    (tmp_path / "a.txt").write_text(
        "Nombre: Test\nEmail: test@test.com\n",
        encoding="utf-8",
    )
    (tmp_path / "b.txt").write_text(
        "Nombre: Other\nEmail: other@test.com\n",
        encoding="utf-8",
    )

    folder_result = process_folder(tmp_path)
    out = tmp_path / "export.xlsx"
    result_path = export_to_excel(folder_result, output_path=str(out))

    assert result_path == out
    assert out.is_file()
    assert out.stat().st_size > 0


def test_export_to_excel_sheets(tmp_path: Path):
    """Should have 'CVs' and 'Resumen' sheets."""
    docs = {
        "cv.txt": {
            "ok": True,
            "fields": {
                "name": "Test", "email": "t@t.com", "phone": "",
                "summary": "", "skills": ["python"], "experience": [],
                "education": [], "languages": [],
            },
            "file_type": "txt",
            "path": "/tmp/cv.txt",
        },
    }

    out = tmp_path / "multi_sheet.xlsx"
    export_to_excel(docs, output_path=str(out))

    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    assert "CVs" in wb.sheetnames
    assert "Resumen" in wb.sheetnames
    wb.close()


# ══════════════════════════════════════════════════════════════════════════════
# DC06 — merge_exports
# ══════════════════════════════════════════════════════════════════════════════


def test_merge_exports_combines_two():
    export1 = {
        "cv1.txt": {
            "ok": True,
            "fields": {"name": "Ana", "email": "ana@test.com", "phone": "555"},
        },
    }
    export2 = {
        "cv2.txt": {
            "ok": True,
            "fields": {"name": "Luis", "email": "luis@test.com", "phone": "123"},
        },
    }

    merged = merge_exports(export1, export2)
    assert len(merged) == 2
    assert "cv1.txt" in merged
    assert "cv2.txt" in merged


def test_merge_exports_deduplicates_by_email():
    export1 = {
        "old.txt": {
            "ok": True,
            "fields": {"name": "Ana", "email": "ana@test.com"},
        },
    }
    export2 = {
        "new.txt": {
            "ok": True,
            "fields": {
                "name": "Ana María",
                "email": "ana@test.com",
                "phone": "555",
                "skills": ["python"],
            },
        },
    }

    merged = merge_exports(export1, export2, dedup_by="email")
    # Should keep the one with more filled fields
    assert len(merged) == 1
    kept = list(merged.values())[0]
    # The more complete one should win (has phone and skills)
    fields = kept.get("fields", kept)
    assert fields.get("phone") == "555" or fields.get("name") == "Ana María"


def test_merge_exports_empty():
    merged = merge_exports({}, {})
    assert merged == {}


def test_merge_exports_single():
    export = {"cv.txt": {"fields": {"name": "Test", "email": "t@t.com"}}}
    merged = merge_exports(export)
    assert len(merged) == 1


def test_merge_exports_no_email_no_duplicate():
    """Entries without email should not be deduped; filename collision avoided."""
    export1 = {
        "cv.txt": {"fields": {"name": "Ana"}},
    }
    export2 = {
        "cv.txt": {"fields": {"name": "Luis"}},
    }

    merged = merge_exports(export1, export2)
    # Both should be present (second gets a renamed key)
    assert len(merged) >= 2


def test_merge_exports_accepts_process_folder_format():
    """merge_exports should normalize process_folder format."""
    export1 = {"results": {"a.txt": {"fields": {"email": "a@a.com"}}}}
    export2 = {"results": {"b.txt": {"fields": {"email": "b@b.com"}}}}

    merged = merge_exports(export1, export2)
    assert len(merged) == 2


# ══════════════════════════════════════════════════════════════════════════════
# DC06 — stats_report
# ══════════════════════════════════════════════════════════════════════════════


def test_stats_report_counts_correctly():
    docs = {
        "cv1.txt": {
            "ok": True,
            "fields": {
                "name": "Ana", "email": "ana@test.com", "phone": "555",
                "skills": ["python", "react", "docker"],
                "languages": ["Inglés", "Español"],
                "experience": ["Dev at X"],
                "education": ["Ingeniería — UCV"],
                "summary": "",
            },
            "file_type": "txt",
            "method": "heuristic",
        },
        "cv2.pdf": {
            "ok": True,
            "fields": {
                "name": "Luis", "email": "luis@test.com", "phone": None,
                "skills": ["python", "java"],
                "languages": ["Inglés"],
                "experience": [],
                "education": [],
                "summary": "",
            },
            "file_type": "pdf",
            "method": "heuristic",
        },
        "cv3.txt": {
            "ok": False,
            "fields": {},
            "file_type": "txt",
            "method": "heuristic",
        },
    }

    stats = stats_report(docs)
    assert stats["total_docs"] == 3
    assert stats["docs_ok"] == 2
    assert stats["success_rate"] == pytest.approx(66.7, abs=0.1)
    assert stats["with_email"] == 2
    assert stats["with_phone"] == 1

    # by_file_type
    assert stats["by_file_type"].get("txt") == 2
    assert stats["by_file_type"].get("pdf") == 1

    # top_skills
    top_skills = dict(stats["top_skills"])
    assert top_skills.get("python") == 2  # appears in both docs

    # top_languages
    top_langs = dict(stats["top_languages"])
    assert top_langs.get("inglés") == 2


def test_stats_report_empty():
    stats = stats_report({})
    assert stats["total_docs"] == 0
    assert stats["success_rate"] == 0.0


def test_stats_report_all_ok():
    docs = {
        "cv.txt": {
            "ok": True,
            "fields": {"name": "Test"},
            "file_type": "txt",
        },
    }
    stats = stats_report(docs)
    assert stats["docs_ok"] == 1
    assert stats["success_rate"] == 100.0


def test_stats_report_with_process_folder_format():
    folder_result = {
        "results": {
            "a.txt": {
                "ok": True,
                "fields": {"name": "A", "email": "a@a.com", "skills": ["python"]},
                "file_type": "txt",
            },
            "b.txt": {
                "ok": False,
                "fields": {},
                "file_type": "txt",
            },
        },
    }

    stats = stats_report(folder_result)
    assert stats["total_docs"] == 2
    assert stats["docs_ok"] == 1


def test_stats_report_by_method():
    docs = {
        "a.txt": {
            "ok": True,
            "fields": {"name": "A"},
            "file_type": "txt",
            "method": "heuristic",
        },
        "b.txt": {
            "ok": True,
            "fields": {"name": "B"},
            "file_type": "txt",
            "method": "llm",
        },
    }
    stats = stats_report(docs)
    assert "by_method" in stats
    assert stats["by_method"].get("heuristic") == 1
    assert stats["by_method"].get("llm") == 1


# ══════════════════════════════════════════════════════════════════════════════
# DC06 — _count_filled_fields helper
# ══════════════════════════════════════════════════════════════════════════════


def test_count_filled_fields():
    assert _count_filled_fields({"name": "A", "email": "a@a.com", "phone": None}) == 2
    assert _count_filled_fields({"name": "", "email": "", "skills": ["python"]}) == 1
    assert _count_filled_fields({}) == 0


# ══════════════════════════════════════════════════════════════════════════════
# DC06 — export_results includes new fields
# ══════════════════════════════════════════════════════════════════════════════


def test_export_results_json_includes_education_and_languages():
    docs = {
        "cv.txt": {
            "ok": True,
            "fields": {
                "name": "Test", "email": "t@t.com", "phone": "",
                "summary": "", "skills": [], "experience": [],
                "education": ["Ing. Sistemas — UCV"],
                "languages": ["Inglés", "Francés"],
            },
            "file_type": "txt",
            "path": "/tmp/cv.txt",
        },
    }
    result = export_results(docs, format="json")
    parsed = json.loads(result)
    assert parsed[0]["education"] == "Ing. Sistemas — UCV"
    assert parsed[0]["languages"] == "Inglés, Francés"
    assert "status" in parsed[0]


def test_export_results_csv_includes_education_and_languages():
    docs = {
        "cv.txt": {
            "ok": True,
            "fields": {
                "name": "Test", "email": "t@t.com", "phone": "",
                "summary": "", "skills": ["a"], "experience": [],
                "education": ["BSc CS"],
                "languages": ["English"],
            },
            "file_type": "txt",
            "path": "/tmp/cv.txt",
        },
    }
    result = export_results(docs, format="csv")
    assert "education" in result
    assert "languages" in result
    assert "BSc CS" in result
    assert "English" in result
    assert "status" in result
