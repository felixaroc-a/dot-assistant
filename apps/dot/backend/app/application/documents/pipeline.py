"""Doc/CV extraction pipeline (FREE-DC01–DC06).

Feature-flagged via DOC_PIPELINE_ENABLED. Heuristic regex extraction by default;
optional LLM enrichment when DOC_PIPELINE_LLM=true and DEEPSEEK_API_KEY is set.

FREE-DC02: .pdf solo si pypdf/pdfminer/PyMuPDF están declarados en requirements.txt;
.docx solo si python-docx está declarado. Sin instalar deps nuevas en runtime.

FREE-DC03: chunk_text / chunk_by_section / chunk_by_paragraph con chunks indexados.
FREE-DC04: structured output con LLM (DeepSeek) si habilitado; fallback heurístico.
           extract_from_url para descargar y extraer de URLs.
           Campos extendidos: education, languages.
FREE-DC05: process_folder — batch recursivo .txt/.md/.docx/.pdf con max_files y logging.
FREE-DC06: export_results — JSON/CSV a string o disco.
           export_to_excel — .xlsx con formato profesional (openpyxl).
           merge_exports — merge + dedup por email.
           stats_report — estadísticas de resumen.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

import httpx

from app.settings import settings

log = logging.getLogger("dot.doc_pipeline")

_BASE_TEXT_SUFFIXES = frozenset({".txt", ".md", ".csv"})
_BATCH_SUFFIXES = frozenset({".txt", ".md", ".docx"})
_REQUIREMENTS_PATH = Path(__file__).resolve().parents[3] / "requirements.txt"

_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
_PHONE_RE = re.compile(
    r"(?:\+?\d{1,3}[\s.-]?)?(?:\(?\d{2,4}\)?[\s.-]?)?\d{3,4}[\s.-]?\d{3,4}(?:[\s.-]?\d{1,4})?"
)
_NAME_LINE_RE = re.compile(
    r"^(?:nombre|name|full name|apellido[s]?)\s*[:\-]\s*(.+)$",
    re.IGNORECASE | re.MULTILINE,
)

# ── DC04: regex heurístico para campos extendidos ──────────────────────────
_SKILL_KEYWORDS = [
    "python", "javascript", "typescript", "java", "c#", "c\\+\\+", "go", "rust",
    "react", "angular", "vue", "node", "django", "fastapi", "flask", "sql",
    "aws", "azure", "gcp", "docker", "kubernetes", "git", "linux",
    "excel", "word", "powerpoint", "power bi", "tableau",
    "liderazgo", "trabajo en equipo", "comunicación", "gestión de proyectos",
    "ventas", "marketing", "contabilidad", "finanzas", "rh",
]
_COMPILED_SKILLS = [
    re.compile(rf"\b{re.escape(s)}\b", re.IGNORECASE) for s in _SKILL_KEYWORDS
]

_SUMMARY_PATTERNS = [
    re.compile(
        r"(?:resumen|summary|perfil|profile|objetivo|objective|about me|acerca de)[\s:]*\n?(.{50,500})",
        re.IGNORECASE | re.DOTALL,
    ),
]

_EXPERIENCE_SECTION_RE = re.compile(
    r"(?:experiencia|experience|experiencia laboral|work experience|empleo|employment|historial laboral)[\s:]*\n?(.+?)(?=\n(?:educación|education|habilidades|skills|certificaciones|idiomas|languages|$)|\Z)",
    re.IGNORECASE | re.DOTALL,
)

_EXPERIENCE_BULLET_RE = re.compile(
    r"(?:^|\n)\s*(?:[-•*]|\d+[.)])\s*(.{20,200})",
    re.MULTILINE,
)

# DC04 extended: education & languages
_EDUCATION_SECTION_RE = re.compile(
    r"(?:educación|education|formación académica|academic background|estudios|studies)[\s:]*\n?(.+?)(?=\n(?:experiencia|experience|habilidades|skills|idiomas|languages|certificaciones|$)|\Z)",
    re.IGNORECASE | re.DOTALL,
)
_EDUCATION_KEYWORDS = [
    "universidad", "universidad", "instituto", "colegio", "licenciatura", "ingeniería",
    "ingenieria", "maestría", "maestria", "master", "doctorado", "phd", "bachiller",
    "técnico", "tecnico", "tecnólogo", "tecnologo", "diplomado", "certificación",
    "certificacion", "bachelor", "degree", "university", "college", "school",
    "mba", "msc", "bsc", "ba ", "bs ",
]
_EDUCATION_BULLET_RE = re.compile(
    r"(?:^|\n)\s*(?:[-•*]|\d+[.)])\s*(.{15,200})",
    re.MULTILINE,
)

_LANGUAGE_NAMES = [
    "inglés", "ingles", "english", "español", "espanol", "spanish",
    "francés", "frances", "french", "alemán", "aleman", "german",
    "portugués", "portugues", "portuguese", "italiano", "italian",
    "chino", "chinese", "mandarín", "mandarin", "japonés", "japones",
    "japanese", "árabe", "arabe", "arabic", "ruso", "russian",
    "holandés", "holandes", "dutch", "coreano", "korean",
]
_LANGUAGE_LINE_RE = re.compile(
    r"(?:idiomas|languages|lenguajes)[\s:]*\n?(.+?)(?:\n\n|\n(?:[A-ZÁÉÍÓÚ][a-záéíóú]+:)|$)",
    re.IGNORECASE | re.DOTALL,
)

# PyMuPDF availability flag for PDF extraction
_PYMUPDF_AVAILABLE: bool | None = None


def _check_pymupdf() -> bool:
    """Check if PyMuPDF (fitz) is available."""
    global _PYMUPDF_AVAILABLE
    if _PYMUPDF_AVAILABLE is None:
        try:
            import fitz  # noqa: F401
            _PYMUPDF_AVAILABLE = True
        except ImportError:
            _PYMUPDF_AVAILABLE = False
    return _PYMUPDF_AVAILABLE


def doc_pipeline_enabled() -> bool:
    return bool(settings.doc_pipeline_enabled)


# ══════════════════════════════════════════════════════════════════════════════
# DC03 — Chunking
# ══════════════════════════════════════════════════════════════════════════════

def chunk_text(
    text: str,
    max_chars: int | None = None,
    overlap: int = 200,
) -> list[dict[str, Any]]:
    """Partir texto largo en chunks con solapamiento.

    Args:
        text: Texto a dividir.
        max_chars: Tamaño máximo por chunk (usa DOC_PIPELINE_CHUNK_SIZE si None).
        overlap: Caracteres de solapamiento entre chunks consecutivos.

    Returns:
        Lista de dicts con keys: index, text, start_char, end_char, char_count.
    """
    if max_chars is None:
        max_chars = settings.doc_pipeline_chunk_size

    raw = (text or "").strip()
    if not raw or len(raw) <= max_chars:
        return [
            {
                "index": 0,
                "text": raw,
                "start_char": 0,
                "end_char": len(raw),
                "char_count": len(raw),
            }
        ]

    chunks: list[dict[str, Any]] = []
    start = 0
    idx = 0

    while start < len(raw):
        end = start + max_chars
        if end >= len(raw):
            chunk_text_val = raw[start:]
        else:
            # Intentar cortar en un límite natural (salto de línea o espacio)
            cut = raw.rfind("\n", start, end)
            if cut > start + max_chars // 2:
                end = cut + 1
            else:
                cut = raw.rfind(" ", start, end)
                if cut > start + max_chars // 2:
                    end = cut + 1

            chunk_text_val = raw[start:end]

        chunk_text_val = chunk_text_val.strip()
        if chunk_text_val:
            chunks.append(
                {
                    "index": idx,
                    "text": chunk_text_val,
                    "start_char": start,
                    "end_char": start + len(chunk_text_val),
                    "char_count": len(chunk_text_val),
                }
            )
            idx += 1

        start = end - overlap if end < len(raw) else end
        if start < 0:
            start = 0

    return chunks


def chunk_by_section(
    text: str,
    max_chars: int | None = None,
    overlap: int = 100,
) -> list[dict[str, Any]]:
    """Split markdown-style text on header boundaries (#, ##, ###).

    Each section starting with a header becomes a chunk. If a single section
    exceeds max_chars, it is further subdivided via chunk_text with overlap.

    Args:
        text: Markdown text to split.
        max_chars: Maximum chars per chunk (uses DOC_PIPELINE_CHUNK_SIZE if None).
        overlap: Overlap chars when subdividing oversized sections.

    Returns:
        Same shape as chunk_text: list of {index, text, start_char, end_char, char_count}.
    """
    if max_chars is None:
        max_chars = settings.doc_pipeline_chunk_size

    raw = (text or "").strip()
    if not raw:
        return [{"index": 0, "text": "", "start_char": 0, "end_char": 0, "char_count": 0}]

    # Detect sections: split on lines starting with # at column 0 or text start
    section_re = re.compile(r"(^|\n)(?=#{1,3}\s+)", re.MULTILINE)
    parts = section_re.split(raw)

    # Reconstruct sections from the split parts
    sections: list[str] = []
    current = ""
    for part in parts:
        if part == "":
            continue
        if re.match(r"^#{1,3}\s+", part):
            if current.strip():
                sections.append(current.strip())
            current = part
        else:
            current += part
    if current.strip():
        sections.append(current.strip())

    if not sections:
        # No headers found — fall back to chunk_text
        return chunk_text(raw, max_chars=max_chars, overlap=overlap)

    chunks: list[dict[str, Any]] = []
    idx = 0
    pos = 0

    for section in sections:
        if len(section) <= max_chars:
            end_pos = pos + len(section)
            chunks.append({
                "index": idx,
                "text": section,
                "start_char": pos,
                "end_char": end_pos,
                "char_count": len(section),
            })
            idx += 1
            pos = end_pos
        else:
            # Subdivide oversized section
            subs = chunk_text(section, max_chars=max_chars, overlap=overlap)
            for sub in subs:
                sub["index"] = idx
                sub["start_char"] = pos + sub["start_char"]
                sub["end_char"] = pos + sub["end_char"]
                idx += 1
            last_sub = subs[-1] if subs else {"end_char": 0}
            pos += last_sub["end_char"]

    return chunks


def chunk_by_paragraph(
    text: str,
    max_chars: int | None = None,
    overlap: int = 100,
) -> list[dict[str, Any]]:
    """Split text on natural paragraph boundaries (double-newline).

    Groups paragraphs into chunks up to max_chars. If a single paragraph
    exceeds max_chars, it is further subdivided via chunk_text with overlap.

    Args:
        text: Text to split.
        max_chars: Maximum chars per chunk (uses DOC_PIPELINE_CHUNK_SIZE if None).
        overlap: Overlap chars when subdividing oversized paragraphs.

    Returns:
        Same shape as chunk_text: list of {index, text, start_char, end_char, char_count}.
    """
    if max_chars is None:
        max_chars = settings.doc_pipeline_chunk_size

    raw = (text or "").strip()
    if not raw:
        return [{"index": 0, "text": "", "start_char": 0, "end_char": 0, "char_count": 0}]

    # Split on double-newline (blank line) boundaries
    paragraphs = re.split(r"\n\s*\n", raw)
    paragraphs = [p.strip() for p in paragraphs if p.strip()]

    if not paragraphs:
        return chunk_text(raw, max_chars=max_chars, overlap=overlap)

    chunks: list[dict[str, Any]] = []
    idx = 0
    pos = 0
    current_group: list[str] = []
    current_len = 0

    for para in paragraphs:
        if current_len + len(para) <= max_chars:
            current_group.append(para)
            current_len += len(para)
        else:
            # Flush current group as a chunk
            if current_group:
                group_text = "\n\n".join(current_group)
                end_pos = pos + len(group_text)
                chunks.append({
                    "index": idx,
                    "text": group_text,
                    "start_char": pos,
                    "end_char": end_pos,
                    "char_count": len(group_text),
                })
                idx += 1
                pos = end_pos

            # Handle oversized single paragraph
            if len(para) > max_chars:
                subs = chunk_text(para, max_chars=max_chars, overlap=overlap)
                for sub in subs:
                    sub["index"] = idx
                    sub["start_char"] = pos + sub["start_char"]
                    sub["end_char"] = pos + sub["end_char"]
                    idx += 1
                last_sub = subs[-1] if subs else {"end_char": 0}
                pos += last_sub["end_char"]
                current_group = []
                current_len = 0
            else:
                current_group = [para]
                current_len = len(para)

    # Flush remaining
    if current_group:
        group_text = "\n\n".join(current_group)
        end_pos = pos + len(group_text)
        chunks.append({
            "index": idx,
            "text": group_text,
            "start_char": pos,
            "end_char": end_pos,
            "char_count": len(group_text),
        })

    return chunks


# ══════════════════════════════════════════════════════════════════════════════
# DC04 — Structured output (LLM + fallback)
# ══════════════════════════════════════════════════════════════════════════════

def _llm_extraction_available() -> bool:
    """Chequea si DOC_PIPELINE_LLM=true y DEEPSEEK_API_KEY está configurada."""
    if not bool(settings.doc_pipeline_llm):
        return False
    return bool(settings.deepseek_api_key)


_STRUCTURED_EXTRACTION_PROMPT = """Eres un extractor de CVs. Del siguiente texto, extrae en JSON válido (solo el JSON, sin markdown):

{
  "name": "Nombre completo o null",
  "email": "email o null",
  "phone": "teléfono o null",
  "summary": "resumen profesional de 1-2 frases o null",
  "skills": ["habilidad1", "habilidad2", ...],
  "experience": ["cargo en empresa — breve descripción", ...],
  "education": ["título — institución", ...],
  "languages": ["idioma - nivel", ...]
}

Reglas:
- Si un campo no aparece, pon null o array vacío.
- No inventes información.
- education: lista de entradas académicas (ej. "Lic. en Computación — Universidad X").
- languages: lista de idiomas con nivel si está presente (ej. "Inglés - Avanzado").
- Responde ÚNICAMENTE con el JSON."""


def _llm_extract_fields(text: str) -> dict[str, Any]:
    """Llama a DeepSeek para extraer campos estructurados del texto."""
    api_key = settings.deepseek_api_key
    url = "https://api.deepseek.com/v1/chat/completions"

    payload = {
        "model": settings.default_chat_model or "deepseek-chat",
        "messages": [
            {"role": "system", "content": _STRUCTURED_EXTRACTION_PROMPT},
            {"role": "user", "content": text[:6000]},  # acotar entrada
        ],
        "temperature": 0.1,
        "max_tokens": 1024,
    }

    try:
        with httpx.Client(timeout=30.0) as client:
            resp = client.post(
                url,
                json=payload,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
            )
            resp.raise_for_status()
            data = resp.json()
            content = data["choices"][0]["message"]["content"].strip()

        # Limpiar fences markdown si el modelo los incluye
        if content.startswith("```"):
            content = re.sub(r"^```(?:json)?\s*", "", content)
            content = re.sub(r"\s*```$", "", content)

        parsed = json.loads(content)
        # Normalizar a la forma esperada
        return {
            "name": parsed.get("name"),
            "email": parsed.get("email"),
            "phone": parsed.get("phone"),
            "summary": parsed.get("summary"),
            "skills": parsed.get("skills") if isinstance(parsed.get("skills"), list) else [],
            "experience": parsed.get("experience") if isinstance(parsed.get("experience"), list) else [],
            "education": parsed.get("education") if isinstance(parsed.get("education"), list) else [],
            "languages": parsed.get("languages") if isinstance(parsed.get("languages"), list) else [],
        }
    except Exception as exc:
        log.warning("LLM extraction failed, falling back to heuristic: %s", exc)
        return _heuristic_extract_fields(text)


def _heuristic_extract_fields(text: str) -> dict[str, Any]:
    """Extracción heurística de campos estructurados extendidos."""
    raw = (text or "").strip()

    # name
    name = _guess_name(raw)

    # email / phone
    emails = _EMAIL_RE.findall(raw)
    phones = [p.strip() for p in _PHONE_RE.findall(raw) if len(re.sub(r"\D", "", p)) >= 7]

    # summary
    summary = None
    for pat in _SUMMARY_PATTERNS:
        m = pat.search(raw)
        if m:
            candidate = re.sub(r"\s+", " ", m.group(1).strip())[:500]
            if len(candidate) > 20:
                summary = candidate
                break

    # skills
    skills: list[str] = []
    for i, skill_re in enumerate(_COMPILED_SKILLS):
        if skill_re.search(raw):
            skills.append(_SKILL_KEYWORDS[i])

    # experience — extract section or bullet points
    experience: list[str] = []
    exp_match = _EXPERIENCE_SECTION_RE.search(raw)
    if exp_match:
        exp_text = exp_match.group(1)
        for bullet in _EXPERIENCE_BULLET_RE.findall(exp_text):
            exp_text_clean = re.sub(r"\s+", " ", bullet.strip())[:200]
            if len(exp_text_clean) > 15 and exp_text_clean not in experience:
                experience.append(exp_text_clean)

    # DC04 extended: education & languages
    education = _extract_education(raw)
    languages = _extract_languages(raw)

    return {
        "name": name,
        "email": emails[0] if emails else None,
        "phone": phones[0] if phones else None,
        "summary": summary,
        "skills": skills,
        "experience": experience,
        "education": education,
        "languages": languages,
    }


def _extract_education(text: str) -> list[str]:
    """Extrae entradas de educación desde el texto usando heurísticas."""
    raw = (text or "").strip()
    results: list[str] = []

    edu_match = _EDUCATION_SECTION_RE.search(raw)
    search_text = edu_match.group(1) if edu_match else raw

    # Try bullet points first
    for bullet in _EDUCATION_BULLET_RE.findall(search_text):
        clean = re.sub(r"\s+", " ", bullet.strip())[:200]
        if len(clean) > 10 and clean not in results:
            results.append(clean)

    # If no bullet points found, try line-by-line for keyword matches
    if not results:
        for line in search_text.splitlines():
            clean = line.strip()
            if not clean or len(clean) < 10:
                continue
            clean_lower = clean.lower()
            for kw in _EDUCATION_KEYWORDS:
                if kw in clean_lower:
                    clean_dedup = re.sub(r"\s+", " ", clean)[:200]
                    if clean_dedup not in results:
                        results.append(clean_dedup)
                    break

    return results[:20]


def _extract_languages(text: str) -> list[str]:
    """Extrae idiomas mencionados en el texto usando heurísticas."""
    raw = (text or "").strip()
    found: list[str] = []

    # First try dedicated languages section
    lang_match = _LANGUAGE_LINE_RE.search(raw)
    search_text = lang_match.group(1) if lang_match else raw

    for lang in _LANGUAGE_NAMES:
        if lang.lower() in search_text.lower():
            if lang not in found:
                found.append(lang.capitalize())

    # If nothing in dedicated section, scan full text
    if not found:
        for lang in _LANGUAGE_NAMES:
            pattern = re.compile(rf"\b{re.escape(lang)}\b", re.IGNORECASE)
            if pattern.search(raw):
                found.append(lang.capitalize())

    return found[:15]


# ══════════════════════════════════════════════════════════════════════════════
# Core extraction (DC01–DC04 integrated)
# ══════════════════════════════════════════════════════════════════════════════

def extract_from_text(text: str, *, use_llm: bool = False) -> dict[str, Any]:
    """Return structured fields extracted from plain text."""
    raw = (text or "").strip()
    if not raw:
        return {
            "ok": False,
            "source": "text",
            "error": "Texto vacío.",
            "fields": {},
        }

    # DC04: structured fields (LLM or heuristic)
    if _llm_extraction_available() or use_llm:
        fields = _llm_extract_fields(raw)
        method = "llm"
    else:
        fields = _heuristic_extract_fields(raw)
        method = "heuristic"

    # Backward-compatible envelope for DC01 callers
    emails = _EMAIL_RE.findall(raw)
    phones = [p.strip() for p in _PHONE_RE.findall(raw) if len(re.sub(r"\D", "", p)) >= 7]

    result_fields: dict[str, Any] = {
        "name": fields.get("name"),
        "email": fields.get("email") or (emails[0] if emails else None),
        "emails": emails,
        "phone": fields.get("phone") or (phones[0] if phones else None),
        "phones": phones,
        "summary": fields.get("summary"),
        "skills": fields.get("skills", []),
        "experience": fields.get("experience", []),
        "education": fields.get("education", []),
        "languages": fields.get("languages", []),
        "raw_preview": raw[:500],
    }

    if use_llm and not _llm_extraction_available():
        result_fields["llm_enrichment"] = _llm_enrich_stub(raw)

    return {
        "ok": True,
        "source": "text",
        "method": method,
        "fields": result_fields,
    }


@lru_cache(maxsize=1)
def _declared_requirement_names() -> frozenset[str]:
    """Nombres normalizados de paquetes en requirements.txt (sin instalar nada)."""
    names: set[str] = set()
    if not _REQUIREMENTS_PATH.is_file():
        return frozenset()

    for raw_line in _REQUIREMENTS_PATH.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("-e ") or line.startswith("-r "):
            continue
        pkg = re.split(r"[<>=!;\[]", line, maxsplit=1)[0].strip().lower()
        if pkg:
            names.add(pkg.replace("_", "-"))
    return frozenset(names)


def _pdf_extraction_available() -> bool:
    declared = _declared_requirement_names()
    if not any(name in declared for name in ("pypdf", "pypdf2", "pdfminer", "pdfminer-six", "pymupdf")):
        return False
    try:
        import pypdf  # noqa: F401
        return True
    except ImportError:
        pass
    try:
        import PyPDF2  # noqa: F401
        return True
    except ImportError:
        pass
    try:
        import pdfminer  # noqa: F401
        return True
    except ImportError:
        pass
    try:
        import fitz  # noqa: F401 -- PyMuPDF
        return True
    except ImportError:
        return False


def _docx_extraction_available() -> bool:
    declared = _declared_requirement_names()
    if "python-docx" not in declared:
        return False
    try:
        import docx  # noqa: F401
        return True
    except ImportError:
        return False


def supported_path_suffixes() -> frozenset[str]:
    """Extensiones soportadas según deps ya declaradas en requirements."""
    suffixes = set(_BASE_TEXT_SUFFIXES)
    if _pdf_extraction_available():
        suffixes.add(".pdf")
    if _docx_extraction_available():
        suffixes.add(".docx")
    return frozenset(suffixes)


def _extract_pdf_text(file_path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore[no-redef]
        except ImportError:
            PdfReader = None  # type: ignore[assignment]

    if PdfReader is not None:
        reader = PdfReader(str(file_path))
        chunks: list[str] = []
        for page in reader.pages:
            text = (page.extract_text() or "").strip()
            if text:
                chunks.append(text)
        if chunks:
            return "\n".join(chunks)

    # Try PyMuPDF next (most reliable)
    if _check_pymupdf():
        import fitz
        doc = fitz.open(str(file_path))
        pages: list[str] = []
        for page in doc:
            text = page.get_text().strip()
            if text:
                pages.append(text)
        doc.close()
        if pages:
            return "\n".join(pages)

    try:
        from pdfminer.high_level import extract_text
    except ImportError as exc:
        raise RuntimeError("PDF sin texto extraíble y pdfminer no disponible.") from exc

    return (extract_text(str(file_path)) or "").strip()


def _extract_docx_text(file_path: Path) -> str:
    from docx import Document

    document = Document(str(file_path))
    chunks = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    return "\n".join(chunks)


def _read_file_text(file_path: Path) -> tuple[str | None, str | None]:
    suffix = file_path.suffix.lower()
    try:
        if suffix in _BASE_TEXT_SUFFIXES:
            return file_path.read_text(encoding="utf-8", errors="replace"), None
        if suffix == ".pdf":
            if not _pdf_extraction_available():
                return None, "Extracción PDF no disponible (falta pypdf/pdfminer en requirements)."
            return _extract_pdf_text(file_path), None
        if suffix == ".docx":
            if not _docx_extraction_available():
                return None, "Extracción DOCX no disponible (falta python-docx en requirements)."
            return _extract_docx_text(file_path), None
    except OSError as exc:
        return None, f"No se pudo leer el archivo: {exc}"
    except Exception as exc:
        return None, f"No se pudo extraer texto: {exc}"
    return None, f"Tipo no soportado: {suffix or 'sin extensión'}"


def extract_from_path(
    path: str | Path,
    *,
    use_llm: bool = False,
    chunk_large: bool = False,
) -> dict[str, Any]:
    """Read a local file and extract structured fields.

    DC03: Si chunk_large=True y el texto > DOC_PIPELINE_CHUNK_SIZE,
    devuelve 'chunks' con lista de pedazos indexados + metadatos.
    """
    file_path = Path(path)
    if not file_path.is_file():
        return {
            "ok": False,
            "source": "path",
            "path": str(file_path),
            "error": "Archivo no encontrado.",
            "fields": {},
        }

    suffix = file_path.suffix.lower()
    allowed = supported_path_suffixes()
    if suffix not in allowed:
        supported = ", ".join(sorted(allowed))
        return {
            "ok": False,
            "source": "path",
            "path": str(file_path),
            "error": (
                f"Tipo no soportado: {suffix or 'sin extensión'}. "
                f"Use {supported} o pase texto directamente."
            ),
            "fields": {},
        }

    content, read_error = _read_file_text(file_path)
    if read_error:
        return {
            "ok": False,
            "source": "path",
            "path": str(file_path),
            "error": read_error,
            "fields": {},
        }

    result = extract_from_text(content or "", use_llm=use_llm)
    result["path"] = str(file_path.resolve())
    result["source"] = "path"
    result["file_type"] = suffix.lstrip(".")
    result["filename"] = file_path.name

    # DC03: chunking
    if chunk_large and content and len(content) > settings.doc_pipeline_chunk_size:
        result["chunks"] = chunk_text(content)

    return result


def extract_from_url(
    url: str,
    *,
    use_llm: bool = False,
    chunk_large: bool = False,
    timeout: float = 30.0,
    user_agent: str = "Nordik-IA-DocPipeline/1.0",
) -> dict[str, Any]:
    """Download a URL and extract structured fields from its content.

    Uses httpx to fetch the URL, then calls extract_from_text on the response body.
    Handles non-200 responses gracefully.

    Args:
        url: The URL to fetch.
        use_llm: Whether to use LLM enrichment (requires DOC_PIPELINE_LLM=true).
        chunk_large: Whether to chunk large documents.
        timeout: HTTP request timeout in seconds.
        user_agent: User-Agent header string.

    Returns:
        Same shape as extract_from_text, plus 'url' and 'status_code' fields.
    """
    try:
        with httpx.Client(timeout=timeout, follow_redirects=True) as client:
            resp = client.get(
                url,
                headers={
                    "User-Agent": user_agent,
                    "Accept": "text/html,text/plain,application/pdf,*/*",
                },
            )
    except httpx.TimeoutException:
        return {
            "ok": False,
            "source": "url",
            "url": url,
            "error": f"Timeout al descargar URL (límite: {timeout}s).",
            "fields": {},
        }
    except httpx.RequestError as exc:
        return {
            "ok": False,
            "source": "url",
            "url": url,
            "error": f"Error al descargar URL: {exc}",
            "fields": {},
        }

    status_code = resp.status_code

    if status_code != 200:
        return {
            "ok": False,
            "source": "url",
            "url": url,
            "status_code": status_code,
            "error": f"Respuesta HTTP {status_code}.",
            "fields": {},
        }

    content_type = resp.headers.get("content-type", "").lower()

    # Try to extract text from HTML using BeautifulSoup
    text_body: str | None = None
    if "text/html" in content_type:
        text_body = _extract_text_from_html(resp.text)
    elif "text/plain" in content_type:
        text_body = resp.text
    elif "application/pdf" in content_type:
        if _check_pymupdf():
            text_body = _extract_pdf_from_bytes(resp.content)
        else:
            return {
                "ok": False,
                "source": "url",
                "url": url,
                "status_code": status_code,
                "error": "Respuesta PDF: PyMuPDF no disponible para extracción.",
                "fields": {},
            }
    else:
        text_body = resp.text  # best-effort

    if not text_body or not text_body.strip():
        return {
            "ok": False,
            "source": "url",
            "url": url,
            "status_code": status_code,
            "error": "Contenido vacío tras extracción.",
            "fields": {},
        }

    result = extract_from_text(text_body, use_llm=use_llm)
    result["source"] = "url"
    result["url"] = url
    result["status_code"] = status_code
    result["content_type"] = content_type

    if chunk_large and len(text_body) > settings.doc_pipeline_chunk_size:
        result["chunks"] = chunk_text(text_body)

    return result


def _extract_text_from_html(html: str) -> str:
    """Extract plain text from HTML using BeautifulSoup."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        # Fallback: strip tags with regex
        clean = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL | re.IGNORECASE)
        clean = re.sub(r"<style[^>]*>.*?</style>", "", clean, flags=re.DOTALL | re.IGNORECASE)
        clean = re.sub(r"<[^>]+>", " ", clean)
        clean = re.sub(r"\s+", " ", clean)
        return clean.strip()

    soup = BeautifulSoup(html, "html.parser")

    # Remove script/style tags
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()

    text = soup.get_text(separator="\n")
    # Collapse whitespace
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return "\n".join(lines)


def _extract_pdf_from_bytes(content: bytes) -> str:
    """Extract text from PDF bytes using PyMuPDF (fitz)."""
    import fitz  # PyMuPDF
    doc = fitz.open(stream=content, filetype="pdf")
    pages: list[str] = []
    for page in doc:
        text = page.get_text().strip()
        if text:
            pages.append(text)
    doc.close()
    return "\n".join(pages)


def run_doc_pipeline_if_enabled(
    *,
    text: str | None = None,
    path: str | Path | None = None,
    use_llm: bool = False,
    chunk_large: bool = False,
) -> dict[str, Any] | None:
    """Optional hook: returns None when DOC_PIPELINE_ENABLED is false."""
    if not doc_pipeline_enabled():
        return None
    if text is not None:
        return extract_from_text(text, use_llm=use_llm)
    if path is not None:
        return extract_from_path(path, use_llm=use_llm, chunk_large=chunk_large)
    return {
        "ok": False,
        "source": "none",
        "error": "Indique text o path.",
        "fields": {},
    }


# ══════════════════════════════════════════════════════════════════════════════
# DC05 — Batch procesar carpeta
# ══════════════════════════════════════════════════════════════════════════════

def process_folder(
    folder_path: str | Path,
    *,
    use_llm: bool = False,
    chunk_large: bool = False,
    recursive: bool = True,
    max_files: int | None = None,
) -> dict[str, Any]:
    """Recorrer carpeta recursivamente y extraer .txt/.md/.docx/.pdf.

    Args:
        folder_path: Ruta a la carpeta.
        use_llm: Si usar enriquecimiento LLM.
        chunk_large: Si particionar documentos grandes.
        recursive: Si recorrer subcarpetas.
        max_files: Límite máximo de archivos a procesar (None = sin límite).

    Returns:
        {
            "ok": True,
            "folder": str,
            "total_files": int,
            "total_processed": int,
            "results": { "filename": { ...extract_from_path result... }, ... },
            "errors": { "filename": "error message", ... },
        }
    """
    folder = Path(folder_path)
    if not folder.is_dir():
        return {
            "ok": False,
            "folder": str(folder),
            "error": "Carpeta no encontrada.",
            "total_files": 0,
            "total_processed": 0,
            "results": {},
            "errors": {},
        }

    results: dict[str, dict[str, Any]] = {}
    errors: dict[str, str] = {}

    glob_method = folder.rglob if recursive else folder.glob

    # Collect files first to apply max_files limit
    batch_suffixes = set(_BATCH_SUFFIXES)
    if _pdf_extraction_available() or _check_pymupdf():
        batch_suffixes.add(".pdf")

    files: list[Path] = []
    for file_path in sorted(glob_method("*")):
        if not file_path.is_file():
            continue
        suffix = file_path.suffix.lower()
        if suffix not in batch_suffixes:
            continue
        if suffix == ".docx" and not _docx_extraction_available():
            continue
        if suffix == ".pdf" and not (_pdf_extraction_available() or _check_pymupdf()):
            continue
        files.append(file_path)

    total_found = len(files)
    if max_files is not None and max_files > 0:
        files = files[:max_files]

    for i, file_path in enumerate(files):
        try:
            result = extract_from_path(file_path, use_llm=use_llm, chunk_large=chunk_large)
            results[file_path.name] = result
        except Exception as exc:
            errors[file_path.name] = str(exc)

        # Progress reporting every 10 files
        if (i + 1) % 10 == 0 or (i + 1) == len(files):
            log.info(
                "doc_pipeline folder progress: %d/%d files processed (%d ok, %d errors)",
                i + 1, len(files), len(results), len(errors),
            )

    total_processed = len(results) + len(errors)
    return {
        "ok": True,
        "folder": str(folder),
        "total_files": total_found,
        "total_processed": total_processed,
        "results": results,
        "errors": errors,
    }


# ══════════════════════════════════════════════════════════════════════════════
# DC06 — Export a JSON/CSV
# ══════════════════════════════════════════════════════════════════════════════

def export_results(
    docs_results: dict[str, Any],
    *,
    format: str = "json",
    output_path: str | Path | None = None,
) -> str | None:
    """Exporta resultados de process_folder (o dict similar) a JSON o CSV.

    Args:
        docs_results: Dict con key "results" → {filename: result_dict} (formato process_folder).
                      También acepta un dict plano de {filename: fields_dict}.
        format: "json" o "csv".
        output_path: Si se provee, escribe a disco y retorna None.
                     Si None, retorna el string generado.

    Returns:
        String JSON/CSV si output_path es None; None si escribe a disco.
    """
    # Normalizar input: aceptar tanto {results: {}} como dict plano
    if isinstance(docs_results, dict) and "results" in docs_results:
        records = docs_results["results"]
    else:
        records = docs_results

    if not isinstance(records, dict):
        raise TypeError("docs_results debe ser un dict con resultados o un dict plano")

    flat_rows: list[dict[str, Any]] = []
    for filename, result in records.items():
        row: dict[str, Any] = {"filename": filename}

        if isinstance(result, dict):
            fields = result.get("fields", {}) if "fields" in result else result
            if isinstance(fields, dict):
                row["name"] = fields.get("name", "")
                row["email"] = fields.get("email", "")
                row["phone"] = fields.get("phone", "")
                row["summary"] = fields.get("summary", "")
                row["skills"] = ", ".join(fields.get("skills", []) or [])
                row["experience"] = " | ".join(fields.get("experience", []) or [])
                row["education"] = " | ".join(fields.get("education", []) or [])
                row["languages"] = ", ".join(fields.get("languages", []) or [])
                row["file_type"] = result.get("file_type", "")
                row["path"] = result.get("path", "")
                row["status"] = "ok" if result.get("ok") else "error"
        flat_rows.append(row)

    if format == "csv":
        return _export_csv(flat_rows, output_path)
    else:
        return _export_json(flat_rows, output_path)


def _export_json(rows: list[dict[str, Any]], output_path: str | Path | None) -> str | None:
    content = json.dumps(rows, ensure_ascii=False, indent=2, default=str)
    if output_path:
        Path(output_path).write_text(content, encoding="utf-8")
        return None
    return content


def _export_csv(rows: list[dict[str, Any]], output_path: str | Path | None) -> str | None:
    if not rows:
        return ""

    # Campos planos — orden estable
    fieldnames = [
        "filename", "name", "email", "phone", "summary",
        "skills", "experience", "education", "languages",
        "file_type", "path", "status",
    ]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)

    content = output.getvalue()
    if output_path:
        Path(output_path).write_text(content, encoding="utf-8", newline="")
        return None
    return content


def export_to_excel(
    docs_results: dict[str, Any],
    *,
    output_path: str | Path,
) -> Path:
    """Exporta resultados a Excel (.xlsx) con formato profesional.

    Usa openpyxl para generar un archivo .xlsx con:
    - Cabeceras congeladas y en negrita.
    - Ancho automático de columnas.
    - Filtros habilitados.
    - Hoja "CVs" con los datos planos.
    - Hoja "Resumen" con estadísticas.

    Args:
        docs_results: Dict con key "results" → {filename: result_dict} (formato process_folder).
                      También acepta un dict plano de {filename: fields_dict}.
        output_path: Ruta donde escribir el .xlsx.

    Returns:
        Path al archivo generado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Normalizar input
    if isinstance(docs_results, dict) and "results" in docs_results:
        records = docs_results["results"]
    else:
        records = docs_results

    if not isinstance(records, dict):
        raise TypeError("docs_results debe ser un dict con resultados o un dict plano")

    wb = Workbook()

    # ── Hoja 1: CVs ──
    ws = wb.active
    ws.title = "CVs"

    headers = [
        "Archivo", "Nombre", "Email", "Teléfono", "Resumen",
        "Habilidades", "Experiencia", "Educación", "Idiomas",
        "Tipo", "Ruta", "Estado",
    ]
    header_keys = [
        "filename", "name", "email", "phone", "summary",
        "skills", "experience", "education", "languages",
        "file_type", "path", "status",
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row_idx = 2
    for filename, result in records.items():
        row: dict[str, Any] = {"filename": filename}
        if isinstance(result, dict):
            fields = result.get("fields", {}) if "fields" in result else result
            if isinstance(fields, dict):
                row["name"] = fields.get("name", "")
                row["email"] = fields.get("email", "")
                row["phone"] = fields.get("phone", "")
                row["summary"] = fields.get("summary", "")
                row["skills"] = ", ".join(fields.get("skills", []) or [])
                row["experience"] = " | ".join(fields.get("experience", []) or [])
                row["education"] = " | ".join(fields.get("education", []) or [])
                row["languages"] = ", ".join(fields.get("languages", []) or [])
                row["file_type"] = result.get("file_type", "")
                row["path"] = result.get("path", "")
                row["status"] = "ok" if result.get("ok") else "error"

        for col_idx, key in enumerate(header_keys, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.border = thin_border
        row_idx += 1

    # Column widths (auto-calculate)
    for col_idx in range(1, len(headers) + 1):
        max_width = len(str(headers[col_idx - 1])) + 2
        for r in range(2, row_idx):
            val = ws.cell(row=r, column=col_idx).value
            if val:
                max_width = max(max_width, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_width + 2

    # Freeze header row + enable autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"

    # ── Hoja 2: Resumen ──
    stats = stats_report(docs_results)
    ws2 = wb.create_sheet("Resumen")

    summary_rows = [
        ["Métrica", "Valor"],
        ["Total documentos", stats.get("total_docs", 0)],
        ["Extracciones exitosas", stats.get("docs_ok", 0)],
        ["Tasa de éxito", f"{stats.get('success_rate', 0):.1f}%"],
        ["Emails encontrados", stats.get("with_email", 0)],
        ["Teléfonos encontrados", stats.get("with_phone", 0)],
        ["", ""],
        ["Por tipo de archivo:", ""],
    ]
    for ft, count in sorted(stats.get("by_file_type", {}).items()):
        summary_rows.append([f"  {ft}", count])
    summary_rows.append(["", ""])
    summary_rows.append(["Top habilidades:", ""])
    for skill_count in stats.get("top_skills", [])[:10]:
        summary_rows.append([f"  {skill_count[0]}", skill_count[1]])
    summary_rows.append(["", ""])
    summary_rows.append(["Top idiomas:", ""])
    for lang_count in stats.get("top_languages", [])[:10]:
        summary_rows.append([f"  {lang_count[0]}", lang_count[1]])

    for r, (label, value) in enumerate(summary_rows, 1):
        cell_a = ws2.cell(row=r, column=1, value=label)
        cell_b = ws2.cell(row=r, column=2, value=value)
        if r == 1:
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_b.font = header_font
            cell_b.fill = header_fill
        elif label and label[0] != " " and label not in ("", "Métrica"):
            cell_a.font = Font(bold=True)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 25

    output = Path(output_path)
    wb.save(str(output))
    return output


def merge_exports(
    *exports: dict[str, Any],
    dedup_by: str = "email",
) -> dict[str, Any]:
    """Merge multiple export dicts into one, deduplicating by a field.

    When duplicate candidates are found, keeps the one with the most fields filled.

    Args:
        *exports: One or more dicts in {filename: result_dict} format.
        dedup_by: Field name to deduplicate by (default "email").

    Returns:
        Merged dict in {filename: result_dict} format.
    """
    merged: dict[str, Any] = {}
    seen: dict[str, str] = {}  # dedup value → filename

    for export in exports:
        # Normalize input
        if isinstance(export, dict) and "results" in export:
            records = export.get("results", {})
        elif isinstance(export, dict):
            records = export
        else:
            continue

        for filename, result in records.items():
            if not isinstance(result, dict):
                merged[filename] = result
                continue

            fields = result.get("fields", {}) if "fields" in result else result
            dedup_val = (fields.get(dedup_by) or "").strip().lower() if isinstance(fields, dict) else ""

            # No dedup value, just add
            if not dedup_val:
                # Avoid filename collision
                base = filename
                counter = 1
                while filename in merged:
                    stem = Path(base).stem
                    suffix = Path(base).suffix
                    filename = f"{stem}_{counter}{suffix}"
                    counter += 1
                merged[filename] = result
                continue

            # Check for duplicates
            if dedup_val in seen:
                existing_name = seen[dedup_val]
                existing = merged[existing_name]

                # Compare filled field count
                existing_fields = existing.get("fields", existing) if isinstance(existing, dict) else {}
                new_filled = _count_filled_fields(fields)
                old_filled = _count_filled_fields(existing_fields)

                if new_filled > old_filled:
                    # Replace old with new (keep same key)
                    merged[existing_name] = result
            else:
                seen[dedup_val] = filename
                merged[filename] = result

    return merged


def _count_filled_fields(fields: dict[str, Any]) -> int:
    """Count non-empty fields in an extraction result."""
    count = 0
    for key, value in (fields or {}).items():
        if value and not (isinstance(value, str) and not value.strip()):
            if isinstance(value, (list, dict)):
                if len(value) > 0:
                    count += 1
            else:
                count += 1
    return count


def stats_report(docs_results: dict[str, Any]) -> dict[str, Any]:
    """Generate summary statistics from document extraction results.

    Args:
        docs_results: Dict with results (from process_folder or similar).

    Returns:
        Dict with keys: total_docs, docs_ok, success_rate, with_email,
        with_phone, by_file_type, top_skills, top_languages, by_method.
    """
    if isinstance(docs_results, dict) and "results" in docs_results:
        records = docs_results["results"]
    elif isinstance(docs_results, dict):
        records = docs_results
    else:
        return {"total_docs": 0, "docs_ok": 0, "success_rate": 0.0}

    total = len(records)
    ok_count = 0
    with_email = 0
    with_phone = 0
    by_file_type: dict[str, int] = {}
    by_method: dict[str, int] = {}
    all_skills: dict[str, int] = {}
    all_languages: dict[str, int] = {}

    for filename, result in records.items():
        if not isinstance(result, dict):
            continue

        is_ok = result.get("ok", True)
        if is_ok:
            ok_count += 1

        fields = result.get("fields", {}) if "fields" in result else result
        if not isinstance(fields, dict):
            continue

        if fields.get("email"):
            with_email += 1
        if fields.get("phone"):
            with_phone += 1

        file_type = result.get("file_type", "unknown")
        by_file_type[file_type] = by_file_type.get(file_type, 0) + 1

        method = result.get("method", "unknown")
        by_method[method] = by_method.get(method, 0) + 1

        for skill in (fields.get("skills") or []):
            skill_lower = skill.lower().strip()
            if skill_lower:
                all_skills[skill_lower] = all_skills.get(skill_lower, 0) + 1

        for lang in (fields.get("languages") or []):
            lang_lower = lang.lower().strip()
            if lang_lower:
                all_languages[lang_lower] = all_languages.get(lang_lower, 0) + 1

    success_rate = (ok_count / total * 100) if total > 0 else 0.0

    top_skills = sorted(all_skills.items(), key=lambda x: x[1], reverse=True)[:20]
    top_languages = sorted(all_languages.items(), key=lambda x: x[1], reverse=True)[:20]

    return {
        "total_docs": total,
        "docs_ok": ok_count,
        "success_rate": round(success_rate, 1),
        "with_email": with_email,
        "with_phone": with_phone,
        "by_file_type": by_file_type,
        "by_method": by_method,
        "top_skills": top_skills,
        "top_languages": top_languages,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Helpers (DC01 legacy)
# ══════════════════════════════════════════════════════════════════════════════

def _guess_name(text: str) -> str | None:
    match = _NAME_LINE_RE.search(text)
    if match:
        return match.group(1).strip()

    for line in text.splitlines():
        candidate = line.strip()
        if not candidate or "@" in candidate:
            continue
        if len(candidate.split()) <= 4 and re.match(r"^[A-Za-zÁÉÍÓÚáéíóúÑñ\s'.-]+$", candidate):
            return candidate
    return None


def _llm_enrich_stub(text: str) -> dict[str, Any]:
    """Placeholder for future LLM enrichment — no API call in stub."""
    return {
        "status": "stub",
        "note": "LLM enrichment pendiente; heurística aplicada.",
        "chars": len(text),
    }
