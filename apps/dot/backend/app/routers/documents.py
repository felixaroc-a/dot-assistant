"""Endpoint para generacion de documentos (Word, Excel, TXT)."""
import logging
import textwrap
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Request
from pydantic import BaseModel, Field

from app.auth_deps import require_product_jwt
from app.dependencies.limiter import limiter
from app.services.document_output_service import (
    add_docx_title_block,
    apply_docx_page_setup,
    get_desktop_work_dir,
    render_markdown_lines_to_docx,
    resolve_output_path,
    sanitize_document_title,
    xlsx_header_styles,
)

log = logging.getLogger("dot.documents")

router = APIRouter(prefix="/v1/documents", tags=["documents"])

DOCUMENT_TYPES = frozenset({"docx", "xlsx", "txt", "pdf"})


class GenerateRequest(BaseModel):
    document_type: str = Field(..., description="Tipo: docx, xlsx, txt")
    title: str = Field(..., min_length=1, max_length=200, description="Nombre del archivo sin extension")
    content: str = Field(..., min_length=1, description="Contenido del documento en markdown/texto plano")
    folder: str | None = Field(default=None, description="Subcarpeta dentro de DOT Trabajos")


class GenerateWithImagesRequest(BaseModel):
    title: str = Field(..., min_length=1, max_length=200, description="Nombre del documento sin extensión")
    content: str = Field(..., min_length=1, description="Contenido markdown. Usa [IMAGE:n] para imágenes")
    image_paths: list[str] = Field(default=[], description="Rutas de imágenes a incrustar")
    folder: str | None = Field(default=None, description="Subcarpeta opcional en DOT Trabajos")


class GenerateWithImagesResponse(BaseModel):
    ok: bool
    filename: str | None = None
    path: str | None = None
    size_bytes: int | None = None
    image_count: int | None = None
    error: str | None = None


class DataSection(BaseModel):
    section_title: str = Field(..., min_length=1, max_length=100, description="Título de la hoja")
    headers: list[str] = Field(..., min_length=1, description="Encabezados de columna")
    rows: list[list[str | int | float]] = Field(..., min_length=1, description="Filas de datos")
    chart_type: str | None = Field(default="bar", description="Tipo de gráfico: bar, line, pie")
    chart_title: str | None = Field(default=None, description="Título del gráfico")


class GenerateSpreadsheetRequest(BaseModel):
    title: str = Field(..., min_length=1, max_length=200, description="Nombre del archivo sin extensión")
    data_sections: list[DataSection] = Field(..., min_length=1, description="Secciones de datos con gráficos")


class GenerateSpreadsheetResponse(BaseModel):
    ok: bool
    filename: str | None = None
    path: str | None = None
    size_bytes: int | None = None
    sheet_count: int | None = None
    error: str | None = None


class GenerateResponse(BaseModel):
    ok: bool
    filename: str
    path: str
    document_type: str
    size_bytes: int


def _generate_docx(title: str, content: str, output_path: Path) -> int:
    """Genera un archivo .docx usando python-docx."""
    from docx import Document

    doc = Document()
    apply_docx_page_setup(doc)
    add_docx_title_block(doc, title)
    render_markdown_lines_to_docx(doc, content)

    doc.core_properties.title = title
    doc.core_properties.author = "DOT IA"

    doc.save(str(output_path))
    return output_path.stat().st_size


def _generate_xlsx(title: str, content: str, output_path: Path) -> int:
    """Genera un archivo .xlsx usando openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_document_title(title)[:31]

    styles = xlsx_header_styles()
    lines = [line for line in content.split("\n") if line.strip()]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(lines[0].split("|")) if lines else 1, 1))
    ws.cell(row=1, column=1, value=title).font = styles["title_font"]
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(lines[0].split("|")) if lines else 1, 1))
    ws.cell(row=2, column=1, value=f"DOT IA — {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = styles["subtitle_font"]

    data_start = 4
    for row_idx, line in enumerate(lines, data_start):
        cells = [c.strip() for c in line.split("|") if c.strip()] if "|" in line else [line]
        for col_idx, cell_value in enumerate(cells, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = styles["thin_border"]
            cell.font = Font(name="Calibri", size=11)
            if row_idx == data_start:
                cell.font = styles["header_font"]
                cell.fill = styles["header_fill"]
                cell.alignment = styles["header_alignment"]

    if lines:
        max_cols = max(len([c.strip() for c in line.split("|") if c.strip()]) if "|" in line else 1 for line in lines)
        for col_idx in range(1, max_cols + 1):
            max_length = 10
            for line in lines:
                cells = [c.strip() for c in line.split("|") if c.strip()] if "|" in line else [line]
                if col_idx - 1 < len(cells):
                    max_length = max(max_length, len(str(cells[col_idx - 1])))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 42)
        ws.freeze_panes = ws.cell(row=data_start + 1, column=1)

    wb.properties.title = title
    wb.properties.creator = "DOT IA"
    wb.save(str(output_path))
    return output_path.stat().st_size


def _generate_txt(title: str, content: str, output_path: Path) -> int:
    """Genera un archivo .txt con UTF-8."""
    header = f"=== {title} ===\n"
    header += f"Generado por DOT -- {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
    header += "=" * 50 + "\n\n"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(header)
        f.write(content)

    return output_path.stat().st_size


def _escape_pdf_text(value: str) -> str:
    return (
        value.replace("\\", "\\\\")
        .replace("(", "\\(")
        .replace(")", "\\)")
    )


def _classify_pdf_line(line: str) -> str:
    stripped = line.strip()
    if stripped.startswith("# "):
        return "h1"
    if stripped.startswith("## "):
        return "h2"
    if stripped.startswith("### "):
        return "h3"
    if stripped.startswith("- ") or stripped.startswith("* "):
        return "bullet"
    return "body"


def _build_pdf_bytes(title: str, content: str) -> bytes:
    """PDF textual con título, encabezados y márgenes legibles (sin deps nativas)."""
    structured: list[tuple[str, str]] = [
        ("title", title),
        ("meta", f"Generado por DOT — {datetime.now().strftime('%d/%m/%Y %H:%M')}"),
        ("spacer", ""),
    ]
    for raw in content.splitlines():
        kind = _classify_pdf_line(raw)
        text = raw.strip()
        if kind == "h1":
            text = text[2:].strip()
        elif kind in {"h2", "h3"}:
            text = text.lstrip("#").strip()
        elif kind == "bullet":
            text = f"• {text[2:].strip()}"
        structured.append((kind, text))

    wrapped: list[tuple[str, str]] = []
    for kind, text in structured:
        if not text:
            wrapped.append((kind, ""))
            continue
        safe = text.encode("latin-1", errors="replace").decode("latin-1")
        width = 78 if kind in {"title", "h1"} else 88
        chunks = textwrap.wrap(safe, width=width) or [""]
        for idx, chunk in enumerate(chunks):
            wrapped.append((kind if idx == 0 else "body", chunk))

    lines_per_page = 40
    pages = [wrapped[i:i + lines_per_page] for i in range(0, len(wrapped), lines_per_page)] or [[("body", "")]]

    objects: list[bytes] = []

    def _add_object(data: bytes) -> int:
        objects.append(data)
        return len(objects)

    _add_object(b"<< /Type /Catalog /Pages 2 0 R >>")
    _add_object(b"<< /Type /Pages /Kids [] /Count 0 >>")

    page_refs: list[int] = []
    font_regular = 3 + len(pages) * 2
    font_bold = font_regular + 1

    for page_lines in pages:
        y = 740
        content_stream = "BT\n"
        for kind, line in page_lines:
            if kind == "spacer" or not line:
                y -= 10
                continue
            if kind == "title":
                size, font, gap = 18, "F2", 24
            elif kind == "meta":
                size, font, gap = 9, "F1", 18
            elif kind == "h1":
                size, font, gap = 14, "F2", 18
            elif kind == "h2":
                size, font, gap = 12, "F2", 14
            elif kind == "h3":
                size, font, gap = 11, "F2", 12
            else:
                size, font, gap = 11, "F1", 14

            content_stream += f"/{font} {size} Tf\n72 {y} Td\n"
            content_stream += f"({_escape_pdf_text(line)}) Tj\n"
            y -= gap

        content_stream += "ET\n"
        stream_bytes = content_stream.encode("latin-1", errors="replace")

        page_obj_num = _add_object(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            f"/Resources << /Font << /F1 {font_regular} 0 R /F2 {font_bold} 0 R >> >> "
            f"/Contents {len(objects) + 1} 0 R >>".encode("ascii")
        )
        page_refs.append(page_obj_num)
        _add_object(
            b"<< /Length " + str(len(stream_bytes)).encode("ascii") + b" >>\nstream\n"
            + stream_bytes
            + b"endstream"
        )

    _add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    _add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")

    kids = " ".join(f"{ref} 0 R" for ref in page_refs)
    objects[1] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_refs)} >>".encode("ascii")

    output = b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output += f"{idx} 0 obj\n".encode("ascii")
        output += obj
        output += b"\nendobj\n"

    xref_start = len(output)
    output += f"xref\n0 {len(objects) + 1}\n".encode("ascii")
    output += b"0000000000 65535 f \n"
    for off in offsets[1:]:
        output += f"{off:010d} 00000 n \n".encode("ascii")
    output += f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n".encode("ascii")
    output += f"startxref\n{xref_start}\n%%EOF\n".encode("ascii")
    return output


def _generate_pdf(title: str, content: str, output_path: Path) -> int:
    """Genera un PDF textual portable sin dependencias nativas."""
    output_path.write_bytes(_build_pdf_bytes(title, content))
    return output_path.stat().st_size


@router.post("/generate", response_model=GenerateResponse)
@limiter.limit("10/minute")
def generate_document(
    request: Request,
    body: GenerateRequest,
    claims: dict = Depends(require_product_jwt),
):
    if body.document_type not in DOCUMENT_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de documento no soportado: {body.document_type}. Usar: {', '.join(sorted(DOCUMENT_TYPES))}",
        )

    get_desktop_work_dir()

    output_path = resolve_output_path(
        kind=body.document_type,
        title=body.title,
        extension=body.document_type,
        folder=body.folder,
    )
    filename = output_path.name

    try:
        if body.document_type == "docx":
            size = _generate_docx(body.title, body.content, output_path)
        elif body.document_type == "xlsx":
            size = _generate_xlsx(body.title, body.content, output_path)
        elif body.document_type == "pdf":
            size = _generate_pdf(body.title, body.content, output_path)
        else:
            size = _generate_txt(body.title, body.content, output_path)
    except ImportError as e:
        log.error("Dependencia faltante: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"Falta dependencia: {e.name}. Instalala con pip install.",
        )
    except Exception:
        log.exception("Error generando documento")
        raise HTTPException(status_code=500, detail="Error al generar el documento.")

    log.info(
        "Documento generado: %s (%d bytes) por usuario %s",
        filename,
        size,
        str(claims.get("sub", "?"))[:8],
    )

    return GenerateResponse(
        ok=True,
        filename=filename,
        path=str(output_path),
        document_type=body.document_type,
        size_bytes=size,
    )


@router.post("/generate-with-images", response_model=GenerateWithImagesResponse)
@limiter.limit("10/minute")
def generate_document_with_images(
    request: Request,
    body: GenerateWithImagesRequest,
    claims: dict = Depends(require_product_jwt),
):
    """Genera un DOCX con imágenes incrustadas usando python-docx."""
    try:
        from app.services.document_image_service import create_docx_with_images

        result = create_docx_with_images(
            title=body.title,
            content=body.content,
            image_paths=body.image_paths if body.image_paths else None,
            folder=body.folder,
        )
        return GenerateWithImagesResponse(**result)
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="python-docx no está instalado.",
        )
    except Exception:
        log.exception("Error generando documento con imágenes")
        raise HTTPException(status_code=500, detail="Error al generar el documento con imágenes.")


@router.post("/generate-spreadsheet", response_model=GenerateSpreadsheetResponse)
@limiter.limit("10/minute")
def generate_spreadsheet(
    request: Request,
    body: GenerateSpreadsheetRequest,
    claims: dict = Depends(require_product_jwt),
):
    """Genera un XLSX con múltiples hojas, datos y gráficos (barras, líneas, torta)."""
    try:
        from app.services.document_image_service import create_xlsx_with_charts

        data_sections = [
            {
                "section_title": s.section_title,
                "headers": s.headers,
                "rows": s.rows,
                "chart_type": s.chart_type or "bar",
                "chart_title": s.chart_title or s.section_title,
            }
            for s in body.data_sections
        ]

        result = create_xlsx_with_charts(
            title=body.title,
            data_sections=data_sections,
        )
        return GenerateSpreadsheetResponse(**result)
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no está instalado.",
        )
    except Exception:
        log.exception("Error generando hoja de cálculo con gráficos")
        raise HTTPException(status_code=500, detail="Error al generar la hoja de cálculo con gráficos.")


# ══════════════════════════════════════════════════════════════════════════════
# Doc/CV extraction pipeline (FREE-DC01–DC06)
# ══════════════════════════════════════════════════════════════════════════════

from app.settings import settings


def _check_doc_pipeline_enabled() -> None:
    """Gate: solo permite acceso si DOC_PIPELINE_ENABLED=true."""
    if not settings.doc_pipeline_enabled:
        raise HTTPException(
            status_code=503,
            detail="Pipeline de documentos deshabilitado. Active DOC_PIPELINE_ENABLED=true en el backend.",
        )


class DocExtractRequest(BaseModel):
    path: str = Field(..., min_length=1, description="Ruta local al archivo (.txt, .md, .docx, .pdf)")
    use_llm: bool = Field(default=False, description="Usar enriquecimiento LLM (DeepSeek) si habilitado")
    chunk_large: bool = Field(default=False, description="Particionar documentos grandes en chunks")


class DocExtractTextRequest(BaseModel):
    text: str = Field(..., min_length=1, description="Texto plano a extraer")
    use_llm: bool = Field(default=False, description="Usar enriquecimiento LLM (DeepSeek) si habilitado")


class DocExtractUrlRequest(BaseModel):
    url: str = Field(..., min_length=1, description="URL a descargar y extraer")
    use_llm: bool = Field(default=False, description="Usar enriquecimiento LLM (DeepSeek) si habilitado")
    chunk_large: bool = Field(default=False, description="Particionar documentos grandes en chunks")


class DocProcessFolderRequest(BaseModel):
    folder_path: str = Field(..., min_length=1, description="Ruta local a la carpeta con documentos")
    recursive: bool = Field(default=True, description="Recorrer subcarpetas recursivamente")
    max_files: int | None = Field(default=None, ge=1, description="Límite máximo de archivos a procesar")
    use_llm: bool = Field(default=False, description="Usar enriquecimiento LLM (DeepSeek) si habilitado")


class DocExtractResponse(BaseModel):
    ok: bool
    source: str | None = None
    path: str | None = None
    url: str | None = None
    file_type: str | None = None
    filename: str | None = None
    method: str | None = None
    fields: dict | None = None
    chunks: list | None = None
    error: str | None = None


class DocProcessFolderResponse(BaseModel):
    ok: bool
    folder: str | None = None
    total_files: int | None = None
    total_processed: int | None = None
    results: dict | None = None
    errors: dict | None = None
    error: str | None = None


@router.post("/extract", response_model=DocExtractResponse)
@limiter.limit("30/minute")
def extract_document(
    request: Request,
    body: DocExtractRequest,
    claims: dict = Depends(require_product_jwt),
):
    """Extrae campos estructurados de un archivo local (CV, documento).

    Soporta .txt, .md, .docx, .pdf según dependencias instaladas.
    Usa heurísticas por defecto; LLM opcional si DOC_PIPELINE_LLM=true.
    """
    _check_doc_pipeline_enabled()

    try:
        from app.application.documents import extract_from_path

        result = extract_from_path(
            path=body.path,
            use_llm=body.use_llm,
            chunk_large=body.chunk_large,
        )
        return DocExtractResponse(**result)
    except Exception:
        log.exception("Error extrayendo documento: %s", body.path)
        raise HTTPException(status_code=500, detail="Error interno al extraer documento.")


@router.post("/extract-text", response_model=DocExtractResponse)
@limiter.limit("30/minute")
def extract_text(
    request: Request,
    body: DocExtractTextRequest,
    claims: dict = Depends(require_product_jwt),
):
    """Extrae campos estructurados de texto plano.

    Útil para procesar contenido ya en memoria sin leer archivos.
    Usa heurísticas por defecto; LLM opcional si DOC_PIPELINE_LLM=true.
    """
    _check_doc_pipeline_enabled()

    try:
        from app.application.documents import extract_from_text

        result = extract_from_text(text=body.text, use_llm=body.use_llm)
        return DocExtractResponse(**result)
    except Exception:
        log.exception("Error extrayendo texto")
        raise HTTPException(status_code=500, detail="Error interno al extraer texto.")


@router.post("/extract-url", response_model=DocExtractResponse)
@limiter.limit("10/minute")
def extract_from_url_endpoint(
    request: Request,
    body: DocExtractUrlRequest,
    claims: dict = Depends(require_product_jwt),
):
    """Descarga una URL y extrae campos estructurados del contenido.

    Soporta HTML (extrae texto con BeautifulSoup), texto plano y PDF.
    Usa heurísticas por defecto; LLM opcional si DOC_PIPELINE_LLM=true.
    """
    _check_doc_pipeline_enabled()

    try:
        from app.application.documents import extract_from_url

        result = extract_from_url(
            url=body.url,
            use_llm=body.use_llm,
            chunk_large=body.chunk_large,
        )
        return DocExtractResponse(**result)
    except Exception:
        log.exception("Error extrayendo URL: %s", body.url)
        raise HTTPException(status_code=500, detail="Error interno al extraer URL.")


@router.post("/process-folder", response_model=DocProcessFolderResponse)
@limiter.limit("5/minute")
def process_folder_endpoint(
    request: Request,
    body: DocProcessFolderRequest,
    claims: dict = Depends(require_product_jwt),
):
    """Procesa en lote todos los documentos de una carpeta.

    Recorre recursivamente .txt, .md, .docx, .pdf y extrae campos
    estructurados de cada uno. Retorna resultados y errores por archivo.
    """
    _check_doc_pipeline_enabled()

    try:
        from app.application.documents import process_folder

        result = process_folder(
            folder_path=body.folder_path,
            recursive=body.recursive,
            max_files=body.max_files,
            use_llm=body.use_llm,
        )
        return DocProcessFolderResponse(**result)
    except Exception:
        log.exception("Error procesando carpeta: %s", body.folder_path)
        raise HTTPException(status_code=500, detail="Error interno al procesar carpeta.")
