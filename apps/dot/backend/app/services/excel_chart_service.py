"""Servicio de generación de Excel con estadísticas y gráficos (P2.3).

Permite crear archivos XLSX con tablas de datos y gráficos usando openpyxl.
Soporta gráficos de barras, líneas y torta (pie) con múltiples hojas.
"""

from __future__ import annotations

import logging
import uuid
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.document_image_service import _get_dot_work_dir, _sanitize_filename

log = logging.getLogger("dot.services.excel_chart")

CHART_TYPES = frozenset({"bar", "line", "pie"})

# Colores para gráficos
CHART_COLORS = [
    "2D5F8B",  # azul oscuro
    "E87040",  # naranja
    "47A86C",  # verde
    "9B59B6",  # violeta
    "E74C3C",  # rojo
    "3498DB",  # celeste
    "F1C40F",  # amarillo
    "1ABC9C",  # turquesa
    "E67E22",  # ámbar
    "95A5A6",  # gris
]


def _create_chart(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    chart_type: str,
    chart_title: str,
    data_start_row: int,
    data_end_row: int,
    data_end_col: int,
    header_row: int,
    chart_colors: list[str] | None = None,
) -> openpyxl.chart._chart.ChartBase:
    """Crea un gráfico openpyxl según el tipo especificado."""
    colors = chart_colors or CHART_COLORS

    if chart_type == "bar":
        chart = BarChart()
    elif chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    else:
        raise ValueError(f"Tipo de gráfico no soportado: {chart_type}")

    chart.title = chart_title or ""
    chart.style = 10

    # Referencia de datos: columna de categorías (A) + columnas de valores (B+)
    cats = Reference(ws, min_col=1, min_row=header_row, max_row=data_end_row)
    data_ref = Reference(ws, min_col=2, max_col=data_end_col, min_row=header_row - 1, max_row=data_end_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    # Colores por serie
    try:
        for idx, series in enumerate(chart.series):
            color = colors[idx % len(colors)]
            series.graphicalProperties.solidFill = color
            if chart_type == "line":
                series.graphicalProperties.line.width = 25000  # 2.5pt en EMUs
            elif chart_type == "bar":
                series.graphicalProperties.solidFill = color
    except Exception:
        # Colores opcionales; no bloquear generación del gráfico
        log.debug("Error aplicando colores de serie en gráfico Excel", exc_info=True)

    # Etiquetas de datos para pie chart
    if chart_type == "pie":
        try:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showCatName = True
        except Exception:
            # Etiquetas opcionales; no bloquear generación del gráfico
            log.debug("Error configurando etiquetas de pie chart en Excel", exc_info=True)
    chart.width = 20  # cm
    chart.height = 12  # cm

    return chart


def create_xlsx_chart(
    *,
    headers: list[str],
    rows: list[list],
    chart_type: str = "bar",
    chart_title: str = "",
    sheet_name: str = "Datos",
) -> dict[str, Any]:
    """Genera un archivo XLSX con una hoja de datos y un gráfico.

    Args:
        headers: Nombres de columnas (la primera es categoría del gráfico).
        rows: Filas de datos. Cada fila es una lista de valores (str|int|float).
        chart_type: "bar", "line" o "pie".
        chart_title: Título del gráfico.
        sheet_name: Nombre de la hoja (máx. 31 caracteres).

    Returns:
        dict con ok, filename, path, size_bytes, sheet_count.
    """
    if chart_type not in CHART_TYPES:
        return {"ok": False, "error": f"Tipo de gráfico no soportado: {chart_type}. Usar: {', '.join(sorted(CHART_TYPES))}"}

    if not headers:
        return {"ok": False, "error": "Se requiere al menos un encabezado."}

    if not rows:
        return {"ok": False, "error": "Se requiere al menos una fila de datos."}

    work_dir = _get_dot_work_dir()
    target_folder = work_dir / "Hojas de calculo"
    safe_sheet = _sanitize_filename(sheet_name)[:31] or "Datos"

    unique_id = uuid.uuid4().hex[:8]
    filename = f"{_sanitize_filename(chart_title or 'Grafico')}_{unique_id}.xlsx"
    filepath = target_folder / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = safe_sheet

    _write_table(ws, headers, rows)

    # Crear gráfico
    data_end_row = len(rows) + 1  # +1 por el header
    data_end_col = len(headers)
    chart = _create_chart(
        ws=ws,
        chart_type=chart_type,
        chart_title=chart_title,
        data_start_row=2,  # primera fila de datos
        data_end_row=data_end_row,
        data_end_col=data_end_col,
        header_row=1,  # fila de encabezados
    )

    # Posicionar gráfico debajo de los datos
    chart_row = data_end_row + 4
    ws.add_chart(chart, f"A{chart_row}")

    wb.save(str(filepath))
    size_bytes = filepath.stat().st_size

    log.info("XLSX con gráfico creado: %s (%d bytes, tipo=%s)", filename, size_bytes, chart_type)

    return {
        "ok": True,
        "filename": filename,
        "path": str(filepath),
        "size_bytes": size_bytes,
        "sheet_count": 1,
        "chart_type": chart_type,
    }


def _write_table(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list[str],
    rows: list[list],
) -> None:
    """Escribe headers y datos en la hoja con formato de tabla."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D2D30", end_color="2D2D30", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(value, (int, float)):
                cell.value = value
            else:
                cell.value = str(value) if value is not None else ""
            cell.alignment = data_alignment
            cell.border = thin_border

    # Auto-ajustar ancho de columnas
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row in rows:
            val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_length = max(max_length, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)
